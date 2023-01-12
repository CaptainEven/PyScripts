# encoding=utf-8

import os
import openpyxl
from tqdm import tqdm


def filter_items(wb_path):
    """
    @param wb_path:
    @return:
    """
    if not os.path.isfile(wb_path):
        print("[Err]: invalid wb path: {:s}"
              .format(wb_path))
        exit(-1)

    wb = openpyxl.load_workbook(wb_path)
    ws = wb.active

    # 显示工作表表名：worksheets会以列表的形式返回当前工作簿里所有的工作表表名：
    sheet_list = wb.worksheets

    filtered_degree = []

    # ----- 遍历表
    print("-> 总共 {:d} 个表.".format(len(sheet_list)))
    for ws_idx, ws in enumerate(sheet_list):
        print("-> 过滤表 {:s}...".format(ws.title))
        col_names = []
        for cell in ws[1]:
            col_names.append(cell.value)
        ws.append(col_names)

        # ---------- 遍历表的每一行
        # ----- 第一次遍历: 学历筛选
        with tqdm(total=ws.max_row) as p_bar:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                degree = row[8]
                # if "硕士" in degree or "学士" in degree:  # "学士", "硕士"
                if "硕士" in degree:  # "学士", "硕士"
                    filtered_degree.append(row)
                p_bar.update()

    print("-> 总共 {:d} 职位 满足 [硕士] 要求."
          .format(len(filtered_degree)))
    for pos_idx, position in enumerate(filtered_degree):
        print("\n: NO.{:d}".format(pos_idx + 1))
        for item_name, item in zip(col_names, position):
            print("{:s}: {}".format(item_name, item))
    print("\n\n\n")

    # ----- 第二次遍历: 专业筛选
    filtered_major = []
    for position in filtered_degree:
        master_major = position[9]      # 硕士专业
        bachelor_major = position[10]   # 本科专业
        if master_major is not None:
            if "不限" in master_major or "材料" in master_major:
                filtered_major.append(position)
        if bachelor_major is not None:
            if "不限" in bachelor_major or "材料" in bachelor_major:
                filtered_major.append(position)

    print("-> 总共 {:d} 职位 满足 [专业] 要求."
          .format(len(filtered_major)))
    for pos_idx, position in enumerate(filtered_major):
        print("\nNO.{:d}".format(pos_idx + 1))
        for item_name, item in zip(col_names, position):
            print("{:s}: {}".format(item_name, item))
    print("\n\n\n")

    # ----- 政治面貌筛选
    filtered_political = []
    for position in filtered_major:
        political_status = position[15]
        if "不限" in political_status or "群众" in political_status:
            filtered_political.append(position)
    print("-> 总共 {:d} 职位 满足 [政治面貌] 要求.\n"
          .format(len(filtered_political)))
    for pos_idx, position in enumerate(filtered_political):
        print("\nNO.{:d}".format(pos_idx + 1))
        for item_name, item in zip(col_names, position):
            print("{:s}: {}".format(item_name, item))


if __name__ == "__main__":
    filter_items(wb_path="e:/2023.xlsx")
